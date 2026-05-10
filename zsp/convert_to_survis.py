import openpyxl
import json
import os
import re

BASE = 'f:/recently work/zsp'

# ---- 1. Read Excel ----
wb = openpyxl.load_workbook(f'{BASE}/文献.xlsx')
ws = wb[wb.sheetnames[0]]

papers = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    papers.append({
        'year': str(row[0]).strip(),
        'title': str(row[1]).strip(),
        'doi': str(row[2]).strip(),
        'reason': str(row[3]).strip() if row[3] else ''
    })

# ---- 2. Tag taxonomy & assignment ----

# Define tag categories
tag_categories = {
    "task": {
        "description": "研究任务 (Research Task)"
    },
    "method": {
        "description": "技术方法 (Methodology)"
    },
    "data_source": {
        "description": "数据来源 (Data Source)"
    },
    "focus": {
        "description": "关注疾病/健康问题 (Health Focus)"
    },
    "pub_type": {
        "description": "文献类型 (Publication Type)"
    }
}

# Define authorized tags
authorized_tags = {
    # task
    "task:surveillance": {"description": "疾病监测 (Disease Surveillance)"},
    "task:early_warning": {"description": "早期预警 (Early Warning)"},
    "task:case_prediction": {"description": "病例预测 (Case Prediction)"},
    "task:health_mention": {"description": "健康提及检测 (Health Mention Detection)"},
    "task:sentiment_analysis": {"description": "情感分析 (Sentiment Analysis)"},
    "task:event_detection": {"description": "事件检测 (Event Detection)"},
    "task:text_classification": {"description": "文本分类 (Text Classification)"},
    "task:infodemic": {"description": "信息疫情分析 (Infodemic Analysis)"},

    # method
    "method:deep_learning": {"description": "深度学习 (Deep Learning)"},
    "method:cnn": {"description": "卷积神经网络 (CNN)"},
    "method:machine_learning": {"description": "机器学习 (Machine Learning)"},
    "method:topic_modeling": {"description": "主题建模 (Topic Modeling)"},

    # data_source
    "data_source:twitter": {"description": "Twitter 数据"},
    "data_source:social_media": {"description": "社交媒体数据 (Social Media)"},

    # focus
    "focus:covid19": {"description": "COVID-19"},
    "focus:meningitis": {"description": "脑膜炎 (Meningitis)"},

    # pub_type
    "pub_type:survey": {"description": "综述 (Survey/Review)"},
    "pub_type:system": {"description": "系统描述 (System Description)"},
    "pub_type:method": {"description": "方法研究 (Methodological Study)"},
}

# ---- 3. Assign tags per paper ----

paper_tags = {
    2: ["task:health_mention", "method:deep_learning", "method:cnn",
        "data_source:twitter", "focus:covid19", "pub_type:method"],
    3: ["task:surveillance", "task:early_warning", "task:sentiment_analysis",
        "data_source:twitter", "focus:covid19", "pub_type:method"],
    4: ["task:text_classification",
        "focus:covid19", "pub_type:method"],
    5: ["task:surveillance", "task:health_mention", "method:deep_learning",
        "data_source:social_media", "pub_type:system"],
    6: ["task:early_warning", "task:event_detection",
        "data_source:social_media", "focus:covid19", "focus:meningitis", "pub_type:method"],
    7: ["task:surveillance",
        "data_source:social_media", "pub_type:survey"],
    8: ["task:early_warning", "task:surveillance", "method:machine_learning",
        "data_source:twitter", "focus:covid19", "pub_type:system"],
    9: ["task:case_prediction",
        "data_source:social_media", "focus:covid19", "pub_type:method"],
    10: ["task:early_warning", "task:case_prediction",
         "data_source:social_media", "focus:covid19", "pub_type:method"],
    11: ["task:case_prediction", "task:infodemic",
         "data_source:social_media", "focus:covid19", "pub_type:method"],
    12: ["task:sentiment_analysis", "method:deep_learning",
         "focus:covid19", "pub_type:method"],
    13: ["task:text_classification", "method:topic_modeling",
         "data_source:twitter", "focus:covid19", "pub_type:method"],
    14: ["task:case_prediction",
         "data_source:twitter", "focus:covid19", "pub_type:method"],
}

# ---- 4. Load authors ----
with open(f'{BASE}/authors_from_crossref.json', 'r', encoding='utf-8') as f:
    authors_map = json.load(f)

# ---- 5. Generate BibTeX ----

def escape_bibtex(s):
    return s.replace('{', '\\{').replace('}', '\\}').replace('&', '\\&')

def make_bib_id(title, year):
    # Create a BibTeX key: FirstAuthorYear + First meaningful word
    first_word = title.split()[0].strip(':,.()').capitalize()
    # Remove non-alphanumeric
    first_word = re.sub(r'[^a-zA-Z0-9]', '', first_word)
    if len(first_word) < 3:
        words = title.split()
        for w in words[1:]:
            clean = re.sub(r'[^a-zA-Z0-9]', '', w.capitalize())
            if len(clean) >= 3:
                first_word = clean
                break
    if not first_word:
        first_word = 'Paper'
    return f"{first_word}{year}"

bib_entries = []
for idx, (row_idx, tags) in enumerate(paper_tags.items(), start=2):
    p = papers[idx - 2]
    bib_id = make_bib_id(p['title'], p['year'])
    title_esc = escape_bibtex(p['title'])
    keywords_str = ', '.join(tags)
    doi_clean = p['doi'].replace('https://doi.org/', '')

    # Get author from CrossRef data
    author = authors_map.get(doi_clean, 'Unknown')
    author_esc = escape_bibtex(author)

    entry = f"""@article{{{bib_id},
  title = {{{title_esc}}},
  author = {{{author_esc}}},
  year = {{{p['year']}}},
  doi = {{{doi_clean}}},
  url = {{{p['doi']}}},
  keywords = {{{keywords_str}}}
}}"""
    bib_entries.append(entry)

bib_content = '\n\n'.join(bib_entries) + '\n'

# Write BibTeX
bib_path = f'{BASE}/survis-master/bib/references.bib'
with open(bib_path, 'w', encoding='utf-8') as f:
    f.write(bib_content)
print(f'Wrote {len(bib_entries)} entries to references.bib')

# ---- 5. Generate tag_categories.js ----
tc_js = "const userDefinedTagCategories = " + json.dumps(tag_categories, indent=2, ensure_ascii=False) + ";"
with open(f'{BASE}/survis-master/src/data/tag_categories.js', 'w', encoding='utf-8') as f:
    f.write(tc_js)
print('Wrote tag_categories.js')

# ---- 6. Generate authorized_tags.js ----
at_js = "const userDefinedAuthorizedTags = " + json.dumps(authorized_tags, indent=2, ensure_ascii=False) + ";"
with open(f'{BASE}/survis-master/src/data/authorized_tags.js', 'w', encoding='utf-8') as f:
    f.write(at_js)
print('Wrote authorized_tags.js')

print('\nDone! Tag summary:')
for cat, info in tag_categories.items():
    count = sum(1 for t in authorized_tags if t.startswith(cat + ':'))
    print(f'  {cat} ({info["description"]}): {count} tags')
